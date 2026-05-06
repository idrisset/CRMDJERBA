"""Export routes: Excel/PDF for clients, apartments, prospects"""
import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from core.database import db, SOFT_DELETE_FILTER
from core.auth import get_current_user

router = APIRouter(prefix="/export", tags=["exports"])


@router.get("/clients/excel")
async def export_clients_excel(current_user: dict = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    headers = ["Reference", "Nom", "Telephone", "Email", "Statut", "Source", "Date creation"]
    ws.append(headers)
    async for c in db.clients.find(SOFT_DELETE_FILTER).sort("created_at", -1):
        ws.append([c.get("reference", ""), c.get("nom", ""), c.get("telephone", ""), c.get("email", ""), c.get("statut", ""), c.get("source", ""), c.get("created_at", "")[:10]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=clients_export.xlsx"})

@router.get("/appartements/excel")
async def export_appartements_excel(current_user: dict = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Appartements"
    headers = ["Bloc", "Lot", "Type", "Etage", "Surface Hab.", "Prix", "Statut"]
    ws.append(headers)
    async for a in db.appartements.find(SOFT_DELETE_FILTER).sort("bloc", 1):
        ws.append([a.get("bloc", ""), a.get("numero_lot", ""), a.get("type_appart", ""), a.get("etage", ""), a.get("surface_habitable", ""), a.get("prix", 0), a.get("statut", "")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=appartements_export.xlsx"})

@router.get("/clients/pdf")
async def export_clients_pdf(current_user: dict = Depends(get_current_user)):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("DJERBA CONSTRUCTION - Liste Clients", styles['Title']), Spacer(1, 12)]
    data = [["Ref", "Nom", "Tel", "Statut", "Date"]]
    async for c in db.clients.find(SOFT_DELETE_FILTER).sort("created_at", -1):
        data.append([c.get("reference", ""), c.get("nom", "")[:25], c.get("telephone", ""), c.get("statut", ""), c.get("created_at", "")[:10]])
    table = Table(data, colWidths=[2*cm, 6*cm, 4*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTSIZE', (0, 0), (-1, -1), 8), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=clients_export.pdf"})

@router.get("/prospects/excel")
async def export_prospects_excel(current_user: dict = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    headers = ["Nom", "Telephone", "Ville", "Type", "Budget Min", "Budget Max", "Source"]
    ws.append(headers)
    async for p in db.prospects.find(SOFT_DELETE_FILTER).sort("created_at", -1):
        ws.append([p.get("nom", ""), p.get("telephone", ""), p.get("ville", ""), p.get("type_logement", ""), p.get("budget_min", ""), p.get("budget_max", ""), p.get("source", "")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=prospects_export.xlsx"})

@router.get("/prospects/pdf")
async def export_prospects_pdf(current_user: dict = Depends(get_current_user)):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("DJERBA CONSTRUCTION - Liste Prospects", styles['Title']), Spacer(1, 12)]
    data = [["Nom", "Tel", "Ville", "Type", "Budget"]]
    async for p in db.prospects.find(SOFT_DELETE_FILTER).sort("created_at", -1):
        budget = f"{p.get('budget_min', 0) or 0:,.0f} - {p.get('budget_max', 0) or 0:,.0f}"
        data.append([p.get("nom", "")[:20], p.get("telephone", ""), p.get("ville", "")[:15], p.get("type_logement", ""), budget])
    table = Table(data, colWidths=[4*cm, 3.5*cm, 3*cm, 3*cm, 4.5*cm])
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTSIZE', (0, 0), (-1, -1), 8), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=prospects_export.pdf"})
