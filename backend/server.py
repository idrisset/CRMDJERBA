from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, WebSocket, WebSocketDisconnect, Response
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import JSONResponse
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import bcrypt
import jwt
import hmac
import hashlib
import asyncio
import resend
import json
import io
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import backup_manager
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend config
resend.api_key = os.environ.get("RESEND_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")

# JWT Config
JWT_ALGORITHM = "HS256"

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

# Password utils
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

# JWT utils
def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=60),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

# ============ ROLES & PERMISSIONS ============
ROLES = {
    "super_admin": {
        "label": "Super Administrateur",
        "level": 3,
        "can_manage_users": True,
        "can_delete": True,
        "can_approve": True,
        "can_modify_sensitive": True,
        "needs_approval": False,
    },
    "admin_limited": {
        "label": "Administrateur Limité",
        "level": 2,
        "can_manage_users": False,
        "can_delete": False,
        "can_approve": False,
        "can_modify_sensitive": False,
        "needs_approval": True,
    },
    "user": {
        "label": "Utilisateur",
        "level": 1,
        "can_manage_users": False,
        "can_delete": False,
        "can_approve": False,
        "can_modify_sensitive": False,
        "needs_approval": True,
    },
}

SENSITIVE_ACTIONS = ["delete_client", "delete_appartement", "delete_prospect", "modify_price", "cancel_reservation", "permanent_delete"]

def get_permissions(role: str) -> dict:
    # Backward compat: old "admin" role maps to super_admin
    if role == "admin":
        role = "super_admin"
    return ROLES.get(role, ROLES["user"])

def require_role(user: dict, min_level: int = 1):
    role = user.get("role", "user")
    perms = get_permissions(role)
    if perms["level"] < min_level:
        raise HTTPException(status_code=403, detail="Permissions insuffisantes")
    return perms

async def generate_client_reference() -> str:
    """Generate next client reference like #001, #002..."""
    last = await db.clients.find_one(
        {"reference": {"$exists": True, "$ne": None}},
        sort=[("reference", -1)]
    )
    if last and last.get("reference"):
        try:
            num = int(last["reference"].replace("#", "")) + 1
        except ValueError:
            num = 1
    else:
        num = 1
    return f"#{num:03d}"

async def check_duplicate_client(nom: str, telephone: str, email: str = None, exclude_id: str = None):
    """Check for potential duplicate clients."""
    conditions = []
    if telephone:
        conditions.append({"telephone": telephone})
        conditions.append({"telephone2": telephone})
    if email:
        conditions.append({"email": {"$regex": f"^{email}$", "$options": "i"}})
    if nom:
        conditions.append({"nom": {"$regex": f"^{nom}$", "$options": "i"}})
    
    if not conditions:
        return []
    
    query = {"$or": conditions, **SOFT_DELETE_FILTER}
    if exclude_id:
        query["_id"] = {"$ne": ObjectId(exclude_id)}
    
    duplicates = []
    async for c in db.clients.find(query):
        reasons = []
        if telephone and (c.get("telephone") == telephone or c.get("telephone2") == telephone):
            reasons.append("telephone")
        if email and c.get("email", "").lower() == email.lower():
            reasons.append("email")
        if nom and c.get("nom", "").lower() == nom.lower():
            reasons.append("nom")
        if reasons:
            duplicates.append({
                "id": str(c["_id"]),
                "reference": c.get("reference", ""),
                "nom": c.get("nom", ""),
                "telephone": c.get("telephone", ""),
                "email": c.get("email", ""),
                "reasons": reasons
            })
    return duplicates

# Auth dependency
async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Non authentifié")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Type de token invalide")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# ============ AUDIT LOG HELPER ============
SOFT_DELETE_FILTER = {"deleted_at": {"$eq": None}}

async def audit_log(user: dict, action: str, entity_type: str, entity_id: str, entity_name: str = "", old_values: dict = None, new_values: dict = None):
    """Log an action to the audit_logs collection (immutable)."""
    doc = {
        "user_id": user.get("_id", ""),
        "user_name": user.get("name", user.get("email", "")),
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entity_name": entity_name,
        "old_values": old_values,
        "new_values": new_values,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(doc)

# Models
class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "user"

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class ClientCreate(BaseModel):
    nom: str
    telephone: str
    telephone2: Optional[str] = None
    email: Optional[str] = None
    salaire: Optional[float] = None
    budget_min: Optional[float] = None
    budget_max: Optional[float] = None
    objectif: Optional[str] = None
    mode_paiement: Optional[str] = None
    etage_souhaite: Optional[str] = None
    situation_familiale: Optional[str] = None
    notes: Optional[str] = None
    statut: str = "nouveau"
    appartement_ids: Optional[list] = None
    force_create: Optional[bool] = False

class ClientUpdate(BaseModel):
    nom: Optional[str] = None
    telephone: Optional[str] = None
    telephone2: Optional[str] = None
    email: Optional[str] = None
    salaire: Optional[float] = None
    budget_min: Optional[float] = None
    budget_max: Optional[float] = None
    objectif: Optional[str] = None
    mode_paiement: Optional[str] = None
    etage_souhaite: Optional[str] = None
    situation_familiale: Optional[str] = None
    notes: Optional[str] = None
    statut: Optional[str] = None
    appartement_ids: Optional[list] = None

class ApprovalRequest(BaseModel):
    action: str
    entity_type: str
    entity_id: str
    entity_name: str
    details: Optional[dict] = None

class ResidenceCreate(BaseModel):
    nom: str
    adresse: Optional[str] = None
    description: Optional[str] = None

class ResidenceUpdate(BaseModel):
    nom: Optional[str] = None
    adresse: Optional[str] = None
    description: Optional[str] = None

# ===== PROSPECTS (Fiche client immobilière - Big Data) =====
class ProspectCreate(BaseModel):
    nom: str
    telephone: str
    telephone2: Optional[str] = None
    email: Optional[str] = None
    ville: Optional[str] = None
    quartier: Optional[str] = None
    type_logement: Optional[str] = None
    etage_souhaite: Optional[str] = None
    nombre_pieces: Optional[int] = None
    budget_min: Optional[float] = None
    budget_max: Optional[float] = None
    mode_paiement: Optional[str] = None
    objectif: Optional[str] = None
    situation_familiale: Optional[str] = None
    notes: Optional[str] = None
    source: Optional[str] = "foire"

class ProspectUpdate(BaseModel):
    nom: Optional[str] = None
    telephone: Optional[str] = None
    telephone2: Optional[str] = None
    email: Optional[str] = None
    ville: Optional[str] = None
    quartier: Optional[str] = None
    type_logement: Optional[str] = None
    etage_souhaite: Optional[str] = None
    nombre_pieces: Optional[int] = None
    budget_min: Optional[float] = None
    budget_max: Optional[float] = None
    mode_paiement: Optional[str] = None
    objectif: Optional[str] = None
    situation_familiale: Optional[str] = None
    notes: Optional[str] = None
    source: Optional[str] = None

class AppartementCreate(BaseModel):
    residence_id: str
    type_appart: str
    prix: float
    etage: str = ""
    statut: str = "disponible"
    surface: Optional[float] = None
    surface_habitable: Optional[float] = None
    surface_utile: Optional[float] = None
    description: Optional[str] = None
    bloc: Optional[str] = None
    numero_lot: Optional[str] = None
    destination: Optional[str] = None

class AppartementUpdate(BaseModel):
    type_appart: Optional[str] = None
    prix: Optional[float] = None
    etage: Optional[str] = None
    statut: Optional[str] = None
    surface: Optional[float] = None
    surface_habitable: Optional[float] = None
    surface_utile: Optional[float] = None
    description: Optional[str] = None
    client_id: Optional[str] = None
    bloc: Optional[str] = None
    numero_lot: Optional[str] = None
    destination: Optional[str] = None

class WhatsAppMessage(BaseModel):
    phone: str
    message: str

class NotificationSettings(BaseModel):
    email_enabled: bool = True
    notification_emails: List[str] = []

# WebSocket Manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager = ConnectionManager()

# Create the app
app = FastAPI(title="DJERBA CONSTRUCTION CRM API")
api_router = APIRouter(prefix="/api")

# CORS - Allow all origins for mobile compatibility
app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ HELPER FUNCTIONS ============
async def send_notification_email(subject: str, html_content: str):
    """Send notification email to all configured recipients"""
    try:
        settings = await db.settings.find_one({"type": "notifications"})
        if not settings or not settings.get("email_enabled"):
            return
        
        recipients = settings.get("notification_emails", [])
        if not recipients:
            return
        
        for email in recipients:
            params = {
                "from": SENDER_EMAIL,
                "to": [email],
                "subject": subject,
                "html": html_content
            }
            await asyncio.to_thread(resend.Emails.send, params)
            logger.info(f"Notification email sent to {email}")
    except Exception as e:
        logger.error(f"Failed to send notification email: {e}")

async def create_lead_from_whatsapp(phone: str, message: str, ai_response: str):
    """Create or update lead from WhatsApp conversation"""
    existing = await db.clients.find_one({"telephone": phone})
    
    if not existing:
        # Create new lead
        lead_doc = {
            "nom": f"Lead WhatsApp {phone[-4:]}",
            "telephone": phone,
            "email": None,
            "salaire": None,
            "situation_familiale": None,
            "notes": f"Premier message: {message}",
            "statut": "nouveau",
            "temperature": "tiède",
            "source": "whatsapp",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": "whatsapp_bot"
        }
        result = await db.clients.insert_one(lead_doc)
        
        # Send notification
        await send_notification_email(
            subject="🔔 Nouveau lead WhatsApp - DJERBA CONSTRUCTION",
            html_content=f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background: #1E3A5F; color: white; padding: 20px; text-align: center;">
                    <h1 style="margin: 0;">DJERBA CONSTRUCTION</h1>
                </div>
                <div style="padding: 20px; background: #f8f9fa;">
                    <h2 style="color: #1E3A5F;">Nouveau lead reçu via WhatsApp</h2>
                    <p><strong>Téléphone:</strong> {phone}</p>
                    <p><strong>Message:</strong> {message}</p>
                    <p><strong>Réponse IA:</strong> {ai_response[:200]}...</p>
                    <a href="{os.environ.get('FRONTEND_URL')}/clients" 
                       style="display: inline-block; background: #C41E3A; color: white; padding: 12px 24px; text-decoration: none; border-radius: 4px; margin-top: 16px;">
                        Voir dans le CRM
                    </a>
                </div>
            </div>
            """
        )
        
        await manager.broadcast({"type": "new_lead", "data": {"phone": phone}})
        return str(result.inserted_id)
    else:
        # Update existing client notes
        await db.clients.update_one(
            {"telephone": phone},
            {"$set": {
                "notes": f"{existing.get('notes', '')}\n\n[{datetime.now().strftime('%d/%m/%Y %H:%M')}] {message}",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return str(existing["_id"])

# ============ AUTH ROUTES ============
@api_router.post("/auth/register")
async def register(user: UserRegister):
    existing = await db.users.find_one({"email": user.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    user_doc = {
        "email": user.email.lower(),
        "password_hash": hash_password(user.password),
        "name": user.name,
        "role": user.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, user.email)
    refresh_token = create_refresh_token(user_id)
    
    return {
        "id": user_id,
        "email": user.email,
        "name": user.name,
        "role": user.role,
        "access_token": access_token,
        "refresh_token": refresh_token
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email.lower()})
    if not user:
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    if not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, user["email"])
    refresh_token = create_refresh_token(user_id)
    
    # Audit login
    user_dict = {"_id": user_id, "name": user.get("name", ""), "email": user["email"]}
    await audit_log(user_dict, "LOGIN", "session", user_id, user.get("name", user["email"]))
    
    return {
        "id": user_id,
        "email": user["email"],
        "name": user.get("name", ""),
        "role": user.get("role", "commercial"),
        "access_token": access_token,
        "refresh_token": refresh_token
    }

@api_router.post("/auth/logout")
async def logout(request: Request):
    # Try to log the logout action if token is valid
    try:
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if token:
            payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
            if user:
                user_dict = {"_id": str(user["_id"]), "name": user.get("name", ""), "email": user.get("email", "")}
                await audit_log(user_dict, "LOGOUT", "session", str(user["_id"]), user.get("name", ""))
    except Exception:
        pass  # Logout should always succeed even with invalid/expired token
    
    response = JSONResponse(content={"message": "Déconnexion réussie"})
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return response

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    role = current_user.get("role", "user")
    perms = get_permissions(role)
    return {
        **{k: v for k, v in current_user.items() if k != "password_hash"},
        "permissions": perms
    }

@api_router.post("/auth/refresh")
async def refresh_token(request: Request):
    refresh = request.cookies.get("refresh_token")
    if not refresh:
        raise HTTPException(status_code=401, detail="Token de rafraîchissement manquant")
    try:
        payload = jwt.decode(refresh, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Type de token invalide")
        
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
        
        access_token = create_access_token(str(user["_id"]), user["email"])
        response = JSONResponse(content={"message": "Token rafraîchi"})
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
        return response
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# ============ CLIENTS ROUTES ============
@api_router.get("/clients")
async def get_clients(current_user: dict = Depends(get_current_user)):
    result = []
    async for c in db.clients.find(SOFT_DELETE_FILTER):
        # Get apartment details for this client
        appart_ids = c.get("appartement_ids") or []
        # Backward compat: old single appartement_id
        if not appart_ids and c.get("appartement_id"):
            appart_ids = [c["appartement_id"]]
        
        appartements_info = []
        for aid in appart_ids:
            try:
                a = await db.appartements.find_one({"_id": ObjectId(aid)})
                if a:
                    appartements_info.append({
                        "id": str(a["_id"]),
                        "numero_lot": a.get("numero_lot", ""),
                        "bloc": a.get("bloc", ""),
                        "type_appart": a.get("type_appart", ""),
                        "statut": a.get("statut", ""),
                    })
            except Exception:
                pass
        
        client_data = {
            "id": str(c["_id"]),
            "reference": c.get("reference", ""),
            "nom": c.get("nom", ""),
            "telephone": c.get("telephone", ""),
            "telephone2": c.get("telephone2", ""),
            "email": c.get("email", ""),
            "salaire": c.get("salaire"),
            "budget_min": c.get("budget_min"),
            "budget_max": c.get("budget_max"),
            "objectif": c.get("objectif", ""),
            "mode_paiement": c.get("mode_paiement", ""),
            "etage_souhaite": c.get("etage_souhaite", ""),
            "situation_familiale": c.get("situation_familiale", ""),
            "notes": c.get("notes", ""),
            "statut": c.get("statut", "nouveau"),
            "appartement_ids": appart_ids,
            "appartements_info": appartements_info,
            "source": c.get("source", "manual"),
            "created_at": c.get("created_at", ""),
            "created_by": c.get("created_by", "")
        }
        result.append(client_data)
    return result

@api_router.post("/clients/check-duplicates")
async def check_duplicates(client: ClientCreate, current_user: dict = Depends(get_current_user)):
    """Check for duplicate clients before creation."""
    duplicates = await check_duplicate_client(client.nom, client.telephone, client.email)
    return {"duplicates": duplicates, "has_duplicates": len(duplicates) > 0}

@api_router.post("/clients")
async def create_client(client: ClientCreate, current_user: dict = Depends(get_current_user)):
    client_data = client.model_dump()
    appartement_ids = client_data.pop("appartement_ids", None) or []
    force_create = client_data.pop("force_create", False)
    
    # Check for duplicates unless force_create
    if not force_create:
        duplicates = await check_duplicate_client(client.nom, client.telephone, client.email)
        if duplicates:
            return {"id": None, "duplicates": duplicates, "needs_confirmation": True}
    
    # Generate auto reference
    reference = await generate_client_reference()
    
    client_doc = {
        **client_data,
        "reference": reference,
        "appartement_ids": appartement_ids,
        "source": "manual",
        "deleted_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["_id"]
    }
    result = await db.clients.insert_one(client_doc)
    client_id = str(result.inserted_id)
    
    await audit_log(current_user, "CREATE", "client", client_id, client_data.get("nom", ""), new_values=client_data)
    
    # Reserve all assigned apartments
    for appart_id in appartement_ids:
        if appart_id:
            try:
                appart = await db.appartements.find_one({"_id": ObjectId(appart_id)})
                if appart and appart.get("statut") == "disponible":
                    await db.appartements.update_one(
                        {"_id": ObjectId(appart_id)},
                        {"$set": {"statut": "réservé", "client_id": client_id}}
                    )
                    await db.reservations.insert_one({
                        "client_id": client_id,
                        "client_nom": client_data.get("nom", ""),
                        "appartement_id": appart_id,
                        "bloc": appart.get("bloc", ""),
                        "numero_lot": appart.get("numero_lot", ""),
                        "type_appart": appart.get("type_appart", ""),
                        "action": "réservé",
                        "agent": current_user.get("name", current_user.get("email", "")),
                        "date": datetime.now(timezone.utc).isoformat()
                    })
                    await manager.broadcast({"type": "appartement_updated", "data": {"id": appart_id}})
            except Exception:
                pass
    
    await manager.broadcast({"type": "client_created", "data": {"id": client_id}})
    return {"id": client_id, "reference": reference, **client_data}

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, client: ClientUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in client.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["_id"]
    
    existing = await db.clients.find_one({"_id": ObjectId(client_id), **SOFT_DELETE_FILTER})
    if not existing:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    old_vals = {k: existing.get(k) for k in update_data.keys() if existing.get(k) != update_data.get(k)}
    
    # Handle multi-apartment changes
    new_appart_ids = update_data.get("appartement_ids")
    old_appart_ids = existing.get("appartement_ids") or []
    if not old_appart_ids and existing.get("appartement_id"):
        old_appart_ids = [existing["appartement_id"]]
    
    if new_appart_ids is not None:
        old_set = set(old_appart_ids)
        new_set = set([a for a in new_appart_ids if a and a != "none"])
        
        # Release removed apartments
        for aid in old_set - new_set:
            try:
                await db.appartements.update_one(
                    {"_id": ObjectId(aid)},
                    {"$set": {"statut": "disponible", "client_id": None}}
                )
                old_a = await db.appartements.find_one({"_id": ObjectId(aid)})
                await db.reservations.insert_one({
                    "client_id": client_id,
                    "client_nom": existing.get("nom", ""),
                    "appartement_id": aid,
                    "bloc": old_a.get("bloc", "") if old_a else "",
                    "numero_lot": old_a.get("numero_lot", "") if old_a else "",
                    "action": "libéré",
                    "agent": current_user.get("name", current_user.get("email", "")),
                    "date": datetime.now(timezone.utc).isoformat()
                })
                await manager.broadcast({"type": "appartement_updated", "data": {"id": aid}})
            except Exception:
                pass
        
        # Reserve new apartments
        for aid in new_set - old_set:
            try:
                appart = await db.appartements.find_one({"_id": ObjectId(aid)})
                if appart:
                    if appart.get("statut") != "disponible" and appart.get("client_id") and appart.get("client_id") != client_id:
                        raise HTTPException(status_code=409, detail=f"Lot {appart.get('numero_lot')} déjà réservé")
                    await db.appartements.update_one(
                        {"_id": ObjectId(aid)},
                        {"$set": {"statut": "réservé", "client_id": client_id}}
                    )
                    await db.reservations.insert_one({
                        "client_id": client_id,
                        "client_nom": update_data.get("nom", existing.get("nom", "")),
                        "appartement_id": aid,
                        "bloc": appart.get("bloc", ""),
                        "numero_lot": appart.get("numero_lot", ""),
                        "type_appart": appart.get("type_appart", ""),
                        "action": "réservé",
                        "agent": current_user.get("name", current_user.get("email", "")),
                        "date": datetime.now(timezone.utc).isoformat()
                    })
                    await manager.broadcast({"type": "appartement_updated", "data": {"id": aid}})
            except HTTPException:
                raise
            except Exception:
                pass
        
        update_data["appartement_ids"] = list(new_set)
        if "appartement_id" in update_data:
            del update_data["appartement_id"]
    
    # Handle status change to vendu
    if client.statut == "vendu":
        final_ids = update_data.get("appartement_ids", old_appart_ids)
        for aid in final_ids:
            try:
                await db.appartements.update_one(
                    {"_id": ObjectId(aid)},
                    {"$set": {"statut": "vendu"}}
                )
                appart = await db.appartements.find_one({"_id": ObjectId(aid)})
                await db.reservations.insert_one({
                    "client_id": client_id,
                    "client_nom": update_data.get("nom", existing.get("nom", "")),
                    "appartement_id": aid,
                    "bloc": appart.get("bloc", "") if appart else "",
                    "numero_lot": appart.get("numero_lot", "") if appart else "",
                    "action": "vendu",
                    "agent": current_user.get("name", current_user.get("email", "")),
                    "date": datetime.now(timezone.utc).isoformat()
                })
                await manager.broadcast({"type": "appartement_updated", "data": {"id": aid}})
            except Exception:
                pass
    
    await db.clients.update_one({"_id": ObjectId(client_id)}, {"$set": update_data})
    await manager.broadcast({"type": "client_updated", "data": {"id": client_id}})
    await audit_log(current_user, "UPDATE", "client", client_id, existing.get("nom", ""), old_values=old_vals, new_values=update_data)
    return {"id": client_id, **update_data}

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, current_user: dict = Depends(get_current_user)):
    perms = get_permissions(current_user.get("role", "user"))
    existing = await db.clients.find_one({"_id": ObjectId(client_id), **SOFT_DELETE_FILTER})
    if not existing:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # If user needs approval for deletion
    if perms["needs_approval"]:
        approval_doc = {
            "requester_id": current_user["_id"],
            "requester_name": current_user.get("name", current_user.get("email", "")),
            "requester_role": current_user.get("role", "user"),
            "action": "delete_client",
            "entity_type": "client",
            "entity_id": client_id,
            "entity_name": existing.get("nom", ""),
            "details": {"reference": existing.get("reference", ""), "telephone": existing.get("telephone", "")},
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.approval_requests.insert_one(approval_doc)
        await audit_log(current_user, "APPROVAL_REQUEST", "client", client_id, existing.get("nom", ""), new_values={"action": "delete_client"})
        await manager.broadcast({"type": "approval_request_created"})
        asyncio.create_task(send_approval_notification_to_admins(approval_doc))
        return {"message": "Demande d'approbation envoyée au Super Administrateur", "approval_required": True}
    
    # Release all apartments
    appart_ids = existing.get("appartement_ids") or []
    if not appart_ids and existing.get("appartement_id"):
        appart_ids = [existing["appartement_id"]]
    for aid in appart_ids:
        try:
            await db.appartements.update_one(
                {"_id": ObjectId(aid)},
                {"$set": {"statut": "disponible", "client_id": None}}
            )
            await manager.broadcast({"type": "appartement_updated", "data": {"id": aid}})
        except Exception:
            pass
    
    await db.clients.update_one(
        {"_id": ObjectId(client_id)},
        {"$set": {
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user["_id"],
            "deleted_by_name": current_user.get("name", current_user.get("email", ""))
        }}
    )
    await audit_log(current_user, "DELETE", "client", client_id, existing.get("nom", ""))
    await manager.broadcast({"type": "client_deleted", "data": {"id": client_id}})
    return {"message": "Client déplacé dans la corbeille"}

# ============ RESIDENCES ROUTES ============
@api_router.get("/residences")
async def get_residences(current_user: dict = Depends(get_current_user)):
    result = []
    async for r in db.residences.find({}):
        result.append({
            "id": str(r["_id"]),
            "nom": r.get("nom", ""),
            "adresse": r.get("adresse", ""),
            "description": r.get("description", "")
        })
    return result

@api_router.post("/residences")
async def create_residence(residence: ResidenceCreate, current_user: dict = Depends(get_current_user)):
    require_role(current_user, min_level=2)
    
    residence_doc = {
        **residence.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.residences.insert_one(residence_doc)
    
    await manager.broadcast({"type": "residence_created", "data": {"id": str(result.inserted_id)}})
    
    return {"id": str(result.inserted_id), **residence.model_dump()}

@api_router.put("/residences/{residence_id}")
async def update_residence(residence_id: str, residence: ResidenceUpdate, current_user: dict = Depends(get_current_user)):
    require_role(current_user, min_level=2)
    
    update_data = {k: v for k, v in residence.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    result = await db.residences.update_one(
        {"_id": ObjectId(residence_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Résidence non trouvée")
    
    await manager.broadcast({"type": "residence_updated", "data": {"id": residence_id}})
    
    return {"id": residence_id, **update_data}

@api_router.delete("/residences/{residence_id}")
async def delete_residence(residence_id: str, current_user: dict = Depends(get_current_user)):
    require_role(current_user, min_level=3)
    
    result = await db.residences.delete_one({"_id": ObjectId(residence_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Résidence non trouvée")
    
    await manager.broadcast({"type": "residence_deleted", "data": {"id": residence_id}})
    
    return {"message": "Résidence supprimée"}

# ============ PROSPECTS ROUTES (Fiche client immobilière) ============
@api_router.get("/prospects")
async def get_prospects(current_user: dict = Depends(get_current_user)):
    result = []
    async for p in db.prospects.find(SOFT_DELETE_FILTER).sort("created_at", -1):
        result.append({
            "id": str(p["_id"]),
            "nom": p.get("nom", ""),
            "telephone": p.get("telephone", ""),
            "telephone2": p.get("telephone2", ""),
            "email": p.get("email", ""),
            "ville": p.get("ville", ""),
            "quartier": p.get("quartier", ""),
            "type_logement": p.get("type_logement", ""),
            "etage_souhaite": p.get("etage_souhaite", ""),
            "nombre_pieces": p.get("nombre_pieces"),
            "budget_min": p.get("budget_min"),
            "budget_max": p.get("budget_max"),
            "mode_paiement": p.get("mode_paiement", ""),
            "objectif": p.get("objectif", ""),
            "situation_familiale": p.get("situation_familiale", ""),
            "notes": p.get("notes", ""),
            "source": p.get("source", "foire"),
            "created_at": p.get("created_at", ""),
        })
    return result

@api_router.post("/prospects")
async def create_prospect(prospect: ProspectCreate, current_user: dict = Depends(get_current_user)):
    doc = {
        **prospect.model_dump(),
        "deleted_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["_id"]
    }
    result = await db.prospects.insert_one(doc)
    await audit_log(current_user, "CREATE", "prospect", str(result.inserted_id), prospect.nom, new_values=prospect.model_dump())
    await manager.broadcast({"type": "prospect_created", "data": {"id": str(result.inserted_id)}})
    return {"id": str(result.inserted_id), **prospect.model_dump()}

@api_router.put("/prospects/{prospect_id}")
async def update_prospect(prospect_id: str, prospect: ProspectUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in prospect.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée")
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    existing_p = await db.prospects.find_one({"_id": ObjectId(prospect_id), **SOFT_DELETE_FILTER})
    if not existing_p:
        raise HTTPException(status_code=404, detail="Prospect non trouvé")
    old_vals = {k: existing_p.get(k) for k in update_data.keys() if existing_p.get(k) != update_data.get(k)}
    result = await db.prospects.update_one({"_id": ObjectId(prospect_id)}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prospect non trouvé")
    await audit_log(current_user, "UPDATE", "prospect", prospect_id, existing_p.get("nom", ""), old_values=old_vals, new_values=update_data)
    await manager.broadcast({"type": "prospect_updated", "data": {"id": prospect_id}})
    return {"id": prospect_id, **update_data}

@api_router.delete("/prospects/{prospect_id}")
async def delete_prospect(prospect_id: str, current_user: dict = Depends(get_current_user)):
    existing = await db.prospects.find_one({"_id": ObjectId(prospect_id), **SOFT_DELETE_FILTER})
    if not existing:
        raise HTTPException(status_code=404, detail="Prospect non trouvé")
    await db.prospects.update_one(
        {"_id": ObjectId(prospect_id)},
        {"$set": {
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user["_id"],
            "deleted_by_name": current_user.get("name", current_user.get("email", ""))
        }}
    )
    await audit_log(current_user, "DELETE", "prospect", prospect_id, existing.get("nom", ""))
    await manager.broadcast({"type": "prospect_deleted", "data": {"id": prospect_id}})
    return {"message": "Prospect déplacé dans la corbeille"}

@api_router.get("/prospects/analytics")
async def get_prospects_analytics(current_user: dict = Depends(get_current_user)):
    total = await db.prospects.count_documents(SOFT_DELETE_FILTER)
    
    # Top villes
    villes_pipeline = [
        {"$match": {"ville": {"$ne": None, "$ne": ""}, "deleted_at": None}},
        {"$group": {"_id": "$ville", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    villes = []
    async for v in db.prospects.aggregate(villes_pipeline):
        villes.append({"name": v["_id"], "count": v["count"]})
    
    # Top quartiers
    quartiers_pipeline = [
        {"$match": {"quartier": {"$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$quartier", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    quartiers = []
    async for q in db.prospects.aggregate(quartiers_pipeline):
        quartiers.append({"name": q["_id"], "count": q["count"]})
    
    # Types de logement
    types_pipeline = [
        {"$match": {"type_logement": {"$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$type_logement", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    types = []
    async for t in db.prospects.aggregate(types_pipeline):
        types.append({"name": t["_id"], "count": t["count"]})
    
    # Objectifs
    objectifs_pipeline = [
        {"$match": {"objectif": {"$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$objectif", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    objectifs = []
    async for o in db.prospects.aggregate(objectifs_pipeline):
        objectifs.append({"name": o["_id"], "count": o["count"]})
    
    # Budget moyen
    budget_pipeline = [
        {"$match": {"budget_max": {"$ne": None, "$gt": 0}}},
        {"$group": {"_id": None, "avg_min": {"$avg": "$budget_min"}, "avg_max": {"$avg": "$budget_max"}}}
    ]
    budget_avg = {"avg_min": 0, "avg_max": 0}
    async for b in db.prospects.aggregate(budget_pipeline):
        budget_avg = {"avg_min": b.get("avg_min", 0) or 0, "avg_max": b.get("avg_max", 0) or 0}
    
    # Mode de paiement
    paiement_pipeline = [
        {"$match": {"mode_paiement": {"$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$mode_paiement", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    paiements = []
    async for p in db.prospects.aggregate(paiement_pipeline):
        paiements.append({"name": p["_id"], "count": p["count"]})

    # Top villes + quartiers combined
    ville_quartier_pipeline = [
        {"$match": {"ville": {"$ne": None, "$ne": ""}, "quartier": {"$ne": None, "$ne": ""}}},
        {"$group": {"_id": {"ville": "$ville", "quartier": "$quartier"}, "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 15}
    ]
    zones = []
    async for z in db.prospects.aggregate(ville_quartier_pipeline):
        zones.append({"ville": z["_id"]["ville"], "quartier": z["_id"]["quartier"], "count": z["count"]})

    return {
        "total": total,
        "top_villes": villes,
        "top_quartiers": quartiers,
        "top_types": types,
        "objectifs": objectifs,
        "budget_avg": budget_avg,
        "modes_paiement": paiements,
        "top_zones": zones
    }

# ============ DASHBOARD ROUTES ============

# ============ APPARTEMENTS ROUTES ============
@api_router.get("/appartements")
async def get_appartements(current_user: dict = Depends(get_current_user)):
    result = []
    async for a in db.appartements.find(SOFT_DELETE_FILTER):
        result.append({
            "id": str(a["_id"]),
            "residence_id": a.get("residence_id", ""),
            "type_appart": a.get("type_appart", ""),
            "prix": a.get("prix", 0),
            "etage": a.get("etage", ""),
            "statut": a.get("statut", "disponible"),
            "surface": a.get("surface"),
            "surface_habitable": a.get("surface_habitable"),
            "surface_utile": a.get("surface_utile"),
            "description": a.get("description", ""),
            "client_id": a.get("client_id"),
            "bloc": a.get("bloc", ""),
            "numero_lot": a.get("numero_lot", ""),
            "destination": a.get("destination", "")
        })
    return result

@api_router.post("/appartements")
async def create_appartement(appart: AppartementCreate, current_user: dict = Depends(get_current_user)):
    appart_doc = {
        **appart.model_dump(),
        "deleted_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["_id"]
    }
    result = await db.appartements.insert_one(appart_doc)
    
    await audit_log(current_user, "CREATE", "appartement", str(result.inserted_id), appart.numero_lot or "", new_values=appart.model_dump())
    await manager.broadcast({"type": "appartement_created", "data": {"id": str(result.inserted_id)}})
    
    return {"id": str(result.inserted_id), **appart.model_dump()}

@api_router.put("/appartements/{appart_id}")
async def update_appartement(appart_id: str, appart: AppartementUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in appart.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    existing = await db.appartements.find_one({"_id": ObjectId(appart_id), **SOFT_DELETE_FILTER})
    if not existing:
        raise HTTPException(status_code=404, detail="Appartement non trouvé")
    
    old_client_id = existing.get("client_id")
    new_client_id = appart.client_id
    old_vals = {k: existing.get(k) for k in update_data.keys() if existing.get(k) != update_data.get(k)}
    
    # Assigning a client to apartment
    if new_client_id and new_client_id != "none" and new_client_id != old_client_id:
        await db.clients.update_one(
            {"_id": ObjectId(new_client_id)},
            {"$set": {"appartement_id": appart_id, "statut": "réservé"}}
        )
        update_data["statut"] = update_data.get("statut", "réservé")
        await db.reservations.insert_one({
            "client_id": new_client_id,
            "appartement_id": appart_id,
            "bloc": existing.get("bloc", ""),
            "numero_lot": existing.get("numero_lot", ""),
            "type_appart": existing.get("type_appart", ""),
            "action": update_data.get("statut", "réservé"),
            "agent": current_user.get("name", current_user.get("email", "")),
            "date": datetime.now(timezone.utc).isoformat()
        })
        # Release old client if different
        if old_client_id and old_client_id != new_client_id:
            await db.clients.update_one(
                {"_id": ObjectId(old_client_id)},
                {"$set": {"appartement_id": None}}
            )
        await manager.broadcast({"type": "client_updated", "data": {"id": new_client_id}})
    elif appart.statut == "disponible" and old_client_id:
        # Releasing apartment
        await db.clients.update_one(
            {"_id": ObjectId(old_client_id)},
            {"$set": {"appartement_id": None}}
        )
        update_data["client_id"] = None
        await db.reservations.insert_one({
            "client_id": old_client_id,
            "appartement_id": appart_id,
            "bloc": existing.get("bloc", ""),
            "numero_lot": existing.get("numero_lot", ""),
            "action": "libéré",
            "agent": current_user.get("name", current_user.get("email", "")),
            "date": datetime.now(timezone.utc).isoformat()
        })
        await manager.broadcast({"type": "client_updated", "data": {"id": old_client_id}})
    
    result = await db.appartements.update_one(
        {"_id": ObjectId(appart_id)},
        {"$set": update_data}
    )
    
    await manager.broadcast({"type": "appartement_updated", "data": {"id": appart_id}})
    await audit_log(current_user, "UPDATE", "appartement", appart_id, existing.get("numero_lot", ""), old_values=old_vals, new_values=update_data)
    return {"id": appart_id, **update_data}

@api_router.delete("/appartements/{appart_id}")
async def delete_appartement(appart_id: str, current_user: dict = Depends(get_current_user)):
    existing = await db.appartements.find_one({"_id": ObjectId(appart_id), **SOFT_DELETE_FILTER})
    if not existing:
        raise HTTPException(status_code=404, detail="Appartement non trouvé")
    await db.appartements.update_one(
        {"_id": ObjectId(appart_id)},
        {"$set": {
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user["_id"],
            "deleted_by_name": current_user.get("name", current_user.get("email", ""))
        }}
    )
    await audit_log(current_user, "DELETE", "appartement", appart_id, existing.get("numero_lot", ""))
    await manager.broadcast({"type": "appartement_deleted", "data": {"id": appart_id}})
    return {"message": "Appartement déplacé dans la corbeille"}

# ============ RESERVATIONS HISTORY ============
@api_router.get("/reservations")
async def get_reservations(current_user: dict = Depends(get_current_user)):
    result = []
    async for r in db.reservations.find({}).sort("date", -1).limit(100):
        result.append({
            "id": str(r["_id"]),
            "client_id": r.get("client_id", ""),
            "client_nom": r.get("client_nom", ""),
            "appartement_id": r.get("appartement_id", ""),
            "bloc": r.get("bloc", ""),
            "numero_lot": r.get("numero_lot", ""),
            "type_appart": r.get("type_appart", ""),
            "action": r.get("action", ""),
            "agent": r.get("agent", ""),
            "date": r.get("date", "")
        })
    return result
@api_router.get("/dashboard")
async def get_dashboard(current_user: dict = Depends(get_current_user)):
    total_clients = await db.clients.count_documents(SOFT_DELETE_FILTER)
    total_appartements = await db.appartements.count_documents(SOFT_DELETE_FILTER)
    apparts_disponibles = await db.appartements.count_documents({"statut": "disponible", **SOFT_DELETE_FILTER})
    apparts_reserves = await db.appartements.count_documents({"statut": "réservé", **SOFT_DELETE_FILTER})
    apparts_vendus = await db.appartements.count_documents({"statut": "vendu", **SOFT_DELETE_FILTER})
    
    clients_nouveau = await db.clients.count_documents({"statut": "nouveau"})
    clients_interesse = await db.clients.count_documents({"statut": "intéressé"})
    clients_visite = await db.clients.count_documents({"statut": "visite"})
    clients_reserve = await db.clients.count_documents({"statut": "réservé"})
    clients_vendu = await db.clients.count_documents({"statut": "vendu"})
    
    # Temperature breakdown
    clients_chaud = await db.clients.count_documents({"temperature": "chaud"})
    clients_tiede = await db.clients.count_documents({"temperature": "tiède"})
    clients_froid = await db.clients.count_documents({"temperature": "froid"})
    
    # WhatsApp leads
    whatsapp_leads = await db.clients.count_documents({"source": "whatsapp"})
    
    recent_clients = []
    async for c in db.clients.find(SOFT_DELETE_FILTER).sort("created_at", -1).limit(5):
        recent_clients.append({
            "id": str(c["_id"]),
            "nom": c.get("nom", ""),
            "statut": c.get("statut", ""),
            "temperature": c.get("temperature", "froid"),
            "source": c.get("source", "manual"),
            "created_at": c.get("created_at", "")
        })
    
    # Logements only stats
    total_logements = await db.appartements.count_documents({"destination": "Logement"})
    logements_disponibles = await db.appartements.count_documents({"destination": "Logement", "statut": "disponible"})
    logements_reserves = await db.appartements.count_documents({"destination": "Logement", "statut": "réservé"})
    logements_vendus = await db.appartements.count_documents({"destination": "Logement", "statut": "vendu"})
    
    # Stats by bloc
    blocs_stats = {}
    for bloc in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        blocs_stats[bloc] = {
            "total": await db.appartements.count_documents({"bloc": bloc, "destination": "Logement"}),
            "disponible": await db.appartements.count_documents({"bloc": bloc, "destination": "Logement", "statut": "disponible"}),
            "reserve": await db.appartements.count_documents({"bloc": bloc, "destination": "Logement", "statut": "réservé"}),
            "vendu": await db.appartements.count_documents({"bloc": bloc, "destination": "Logement", "statut": "vendu"}),
        }
    
    return {
        "total_clients": total_clients,
        "total_appartements": total_appartements,
        "total_logements": total_logements,
        "logements_disponibles": logements_disponibles,
        "logements_reserves": logements_reserves,
        "logements_vendus": logements_vendus,
        "appartements_disponibles": apparts_disponibles,
        "appartements_reserves": apparts_reserves,
        "appartements_vendus": apparts_vendus,
        "clients_par_statut": {
            "nouveau": clients_nouveau,
            "intéressé": clients_interesse,
            "visite": clients_visite,
            "réservé": clients_reserve,
            "vendu": clients_vendu
        },
        "clients_par_temperature": {
            "chaud": clients_chaud,
            "tiède": clients_tiede,
            "froid": clients_froid
        },
        "whatsapp_leads": whatsapp_leads,
        "recent_clients": recent_clients,
        "blocs_stats": blocs_stats
    }

# ============ WHATSAPP WEBHOOK (Meta Business API) ============
@api_router.get("/whatsapp/webhook")
async def verify_whatsapp_webhook(request: Request):
    """Verify webhook for Meta WhatsApp Business API"""
    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")
    
    verify_token = os.environ.get("WHATSAPP_VERIFY_TOKEN", "")
    
    if mode == "subscribe" and token == verify_token:
        logger.info("WhatsApp webhook verified")
        return Response(content=challenge, status_code=200)
    
    logger.warning("WhatsApp webhook verification failed")
    raise HTTPException(status_code=403, detail="Verification failed")

@api_router.post("/whatsapp/webhook")
async def handle_whatsapp_webhook(request: Request):
    """Handle incoming WhatsApp messages from Meta Business API"""
    body = await request.body()
    
    # Verify signature (if APP_SECRET is configured)
    app_secret = os.environ.get("META_APP_SECRET", "")
    if app_secret:
        signature = request.headers.get("X-Hub-Signature-256", "")
        if signature.startswith("sha256="):
            expected_signature = signature[7:]
            computed_hash = hmac.new(
                app_secret.encode('utf-8'),
                body,
                hashlib.sha256
            ).hexdigest()
            if not hmac.compare_digest(computed_hash, expected_signature):
                logger.error("Invalid WhatsApp webhook signature")
                raise HTTPException(status_code=403, detail="Invalid signature")
    
    try:
        payload = json.loads(body)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    
    if payload.get("object") != "whatsapp_business_account":
        return JSONResponse(status_code=200, content={"status": "ok"})
    
    # Process messages asynchronously
    asyncio.create_task(process_whatsapp_webhook(payload))
    
    return JSONResponse(status_code=200, content={"status": "ok"})

async def process_whatsapp_webhook(payload: dict):
    """Process incoming WhatsApp messages"""
    try:
        for entry in payload.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})
                
                if "messages" in value:
                    for message in value.get("messages", []):
                        await handle_incoming_whatsapp_message(message)
    except Exception as e:
        logger.error(f"WhatsApp webhook processing error: {e}")

async def handle_incoming_whatsapp_message(message: dict):
    """Handle a single incoming WhatsApp message"""
    message_id = message.get("id")
    sender = message.get("from")
    message_type = message.get("type", "text")
    
    if message_type != "text":
        return
    
    message_content = message.get("text", {}).get("body", "")
    
    logger.info(f"WhatsApp message from {sender}: {message_content}")
    
    # Generate AI response
    ai_response = await generate_whatsapp_ai_response(sender, message_content)
    
    # Store conversation
    await db.whatsapp_conversations.insert_one({
        "message_id": message_id,
        "phone": sender,
        "user_message": message_content,
        "ai_response": ai_response,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Create or update lead
    await create_lead_from_whatsapp(sender, message_content, ai_response)
    
    # Send response via WhatsApp API (if configured)
    await send_whatsapp_message(sender, ai_response)

async def generate_whatsapp_ai_response(phone: str, message: str) -> str:
    """Generate AI response for WhatsApp message"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        # Get residences and apartments for context
        residences = []
        async for r in db.residences.find({}):
            residences.append(r.get("nom", ""))
        
        appartements = []
        async for a in db.appartements.find({"statut": "disponible", **SOFT_DELETE_FILTER}):
            appartements.append({
                "type": a.get("type_appart", ""),
                "prix": a.get("prix", 0),
                "etage": a.get("etage", 0),
                "surface": a.get("surface", 0)
            })
        
        system_message = f"""Tu es l'assistant virtuel de DJERBA CONSTRUCTION, une entreprise immobilière de qualité.
Tu réponds en français, en arabe ou en anglais selon la langue du client.
Tu dois:
1. Répondre aux questions sur les appartements disponibles
2. Poser des questions pour qualifier le prospect (budget, type recherché, localisation)
3. Collecter les informations de contact (nom, email si possible)
4. Être professionnel et accueillant

Résidences disponibles: {', '.join(residences) if residences else 'Nos résidences premium'}
Appartements disponibles: {len(appartements)}
Types: {', '.join(set(a['type'] for a in appartements)) if appartements else 'F2, F3, F4'}
Prix: {min(a['prix'] for a in appartements) if appartements else 0:,.0f} DA - {max(a['prix'] for a in appartements) if appartements else 0:,.0f} DA

Sois concis et professionnel. Maximum 2-3 phrases par réponse.
"""
        
        chat = LlmChat(
            api_key=os.environ.get("EMERGENT_LLM_KEY"),
            session_id=f"whatsapp_{phone}",
            system_message=system_message
        ).with_model("openai", "gpt-5.2")
        
        user_message = UserMessage(text=message)
        response = await chat.send_message(user_message)
        
        return response
    except Exception as e:
        logger.error(f"WhatsApp AI error: {e}")
        return "Merci pour votre message. Un de nos conseillers vous contactera bientôt. للتواصل: +213770481500"

async def send_whatsapp_message(recipient: str, message: str):
    """Send WhatsApp message via Meta Business API"""
    import requests
    
    phone_number_id = os.environ.get("WHATSAPP_PHONE_NUMBER_ID", "")
    access_token = os.environ.get("META_ACCESS_TOKEN", "")
    
    if not phone_number_id or not access_token:
        logger.warning("WhatsApp API not configured - message not sent")
        return
    
    url = f"https://graph.facebook.com/v18.0/{phone_number_id}/messages"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": recipient,
        "type": "text",
        "text": {"body": message}
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        logger.info(f"WhatsApp message sent to {recipient}")
    except Exception as e:
        logger.error(f"Failed to send WhatsApp message: {e}")

# ============ WHATSAPP TEST ROUTES ============
@api_router.post("/whatsapp/message")
async def test_whatsapp_message(msg: WhatsAppMessage, current_user: dict = Depends(get_current_user)):
    """Test WhatsApp AI response"""
    try:
        ai_response = await generate_whatsapp_ai_response(msg.phone, msg.message)
        
        await db.whatsapp_conversations.insert_one({
            "phone": msg.phone,
            "user_message": msg.message,
            "ai_response": ai_response,
            "test": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"response": ai_response, "phone": msg.phone}
    except Exception as e:
        logger.error(f"WhatsApp AI error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur IA: {str(e)}")

@api_router.get("/whatsapp/conversations")
async def get_whatsapp_conversations(current_user: dict = Depends(get_current_user)):
    result = []
    async for c in db.whatsapp_conversations.find({}).sort("created_at", -1).limit(50):
        result.append({
            "id": str(c["_id"]),
            "phone": c.get("phone", ""),
            "user_message": c.get("user_message", ""),
            "ai_response": c.get("ai_response", ""),
            "created_at": c.get("created_at", "")
        })
    return result

# ============ SETTINGS ROUTES ============
@api_router.get("/settings/notifications")
async def get_notification_settings(current_user: dict = Depends(get_current_user)):
    require_role(current_user, min_level=2)
    
    settings = await db.settings.find_one({"type": "notifications"})
    if not settings:
        return {"email_enabled": False, "notification_emails": []}
    
    return {
        "email_enabled": settings.get("email_enabled", False),
        "notification_emails": settings.get("notification_emails", [])
    }

@api_router.put("/settings/notifications")
async def update_notification_settings(settings: NotificationSettings, current_user: dict = Depends(get_current_user)):
    require_role(current_user, min_level=3)
    
    await db.settings.update_one(
        {"type": "notifications"},
        {"$set": {
            "type": "notifications",
            "email_enabled": settings.email_enabled,
            "notification_emails": settings.notification_emails,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"message": "Paramètres mis à jour"}

# ============ EXPORT ROUTES ============
@api_router.get("/export/clients/excel")
async def export_clients_excel(current_user: dict = Depends(get_current_user)):
    """Export clients to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Headers
    headers = ["Nom", "Téléphone 1", "Téléphone 2", "Email", "Objectif", "Mode Paiement", "Salaire", "Budget Min", "Budget Max", "Etage Souhaité", "Situation", "Statut", "Source", "Date création"]
    ws.append(headers)
    
    # Data
    async for c in db.clients.find(SOFT_DELETE_FILTER):
        ws.append([
            c.get("nom", ""),
            c.get("telephone", ""),
            c.get("telephone2", ""),
            c.get("email", ""),
            c.get("objectif", ""),
            c.get("mode_paiement", ""),
            c.get("salaire", ""),
            c.get("budget_min", ""),
            c.get("budget_max", ""),
            c.get("etage_souhaite", ""),
            c.get("situation_familiale", ""),
            c.get("statut", ""),
            c.get("source", "manual"),
            c.get("created_at", "")
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=clients_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/export/appartements/excel")
async def export_appartements_excel(current_user: dict = Depends(get_current_user)):
    """Export apartments to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Appartements"
    
    # Get residences for mapping
    residences = {}
    async for r in db.residences.find({}):
        residences[str(r["_id"])] = r.get("nom", "")
    
    # Headers
    headers = ["Lot", "Bloc", "Résidence", "Type", "Étage", "Surface Hab. (m²)", "Surface Utile (m²)", "Prix (DA)", "Statut", "Destination"]
    ws.append(headers)
    
    # Data
    async for a in db.appartements.find(SOFT_DELETE_FILTER):
        ws.append([
            a.get("numero_lot", ""),
            a.get("bloc", ""),
            residences.get(a.get("residence_id", ""), ""),
            a.get("type_appart", ""),
            a.get("etage", ""),
            a.get("surface_habitable", ""),
            a.get("surface_utile", ""),
            a.get("prix", 0),
            a.get("statut", ""),
            a.get("destination", "")
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=appartements_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/export/clients/pdf")
async def export_clients_pdf(current_user: dict = Depends(get_current_user)):
    """Export clients to PDF"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1E3A5F'))
    elements.append(Paragraph("DJERBA CONSTRUCTION - Liste des Clients", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Table data
    data = [["Nom", "Téléphone", "Statut", "Température"]]
    
    async for c in db.clients.find(SOFT_DELETE_FILTER):
        data.append([
            c.get("nom", "")[:20],
            c.get("telephone", ""),
            c.get("statut", ""),
            c.get("temperature", "")
        ])
    
    if len(data) > 1:
        table = Table(data, colWidths=[6*cm, 4*cm, 3*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        elements.append(table)
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=clients_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

# ============ PROSPECTS EXPORT ROUTES ============
@api_router.get("/export/prospects/excel")
async def export_prospects_excel(current_user: dict = Depends(get_current_user)):
    """Export prospects to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    
    headers = ["Nom", "Téléphone 1", "Téléphone 2", "Email", "Ville", "Quartier", "Type Logement", "Étage Souhaité", "Nb Pièces", "Budget Min", "Budget Max", "Mode Paiement", "Objectif", "Situation", "Source", "Notes", "Date création"]
    ws.append(headers)
    
    async for p in db.prospects.find(SOFT_DELETE_FILTER):
        ws.append([
            p.get("nom", ""),
            p.get("telephone", ""),
            p.get("telephone2", ""),
            p.get("email", ""),
            p.get("ville", ""),
            p.get("quartier", ""),
            p.get("type_logement", ""),
            p.get("etage_souhaite", ""),
            p.get("nombre_pieces", ""),
            p.get("budget_min", ""),
            p.get("budget_max", ""),
            p.get("mode_paiement", ""),
            p.get("objectif", ""),
            p.get("situation_familiale", ""),
            p.get("source", ""),
            p.get("notes", ""),
            p.get("created_at", "")
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=prospects_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/export/prospects/pdf")
async def export_prospects_pdf(current_user: dict = Depends(get_current_user)):
    """Export prospects to PDF"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor('#1E3A5F'))
    elements.append(Paragraph("EDIMCO - Prospects Big Data", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    data = [["Nom", "Téléphone", "Ville", "Quartier", "Type", "Source"]]
    async for p in db.prospects.find(SOFT_DELETE_FILTER).sort("created_at", -1):
        data.append([
            p.get("nom", "")[:20],
            p.get("telephone", ""),
            p.get("ville", ""),
            p.get("quartier", "")[:15],
            p.get("type_logement", ""),
            p.get("source", "")
        ])
    
    if len(data) > 1:
        table = Table(data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 2*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        elements.append(table)
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=prospects_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

# ============ USERS ROUTES (Admin only) ============
@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    require_role(current_user, min_level=3)
    
    result = []
    async for u in db.users.find({}, {"password_hash": 0}):
        result.append({
            "id": str(u["_id"]),
            "email": u.get("email", ""),
            "name": u.get("name", ""),
            "role": u.get("role", "user"),
            "is_active": u.get("is_active", True),
            "created_at": u.get("created_at", "")
        })
    return result

@api_router.post("/users")
async def create_user(user: UserRegister, current_user: dict = Depends(get_current_user)):
    """Create a new user - Super Admin only"""
    require_role(current_user, min_level=3)
    
    if user.role not in ROLES:
        raise HTTPException(status_code=400, detail="Rôle invalide")
    
    existing = await db.users.find_one({"email": user.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    user_doc = {
        "email": user.email.lower(),
        "password_hash": hash_password(user.password),
        "name": user.name,
        "role": user.role,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(user_doc)
    await audit_log(current_user, "CREATE", "user", str(result.inserted_id), user.name, new_values={"email": user.email, "role": user.role})
    return {"id": str(result.inserted_id), "email": user.email, "name": user.name, "role": user.role}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, user_update: UserUpdate, current_user: dict = Depends(get_current_user)):
    """Update a user - Super Admin only"""
    require_role(current_user, min_level=3)
    
    update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée")
    
    if "role" in update_data and update_data["role"] not in ROLES:
        raise HTTPException(status_code=400, detail="Rôle invalide")
    
    existing = await db.users.find_one({"_id": ObjectId(user_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update_data})
    await audit_log(current_user, "UPDATE", "user", user_id, existing.get("name", ""), new_values=update_data)
    return {"id": user_id, **update_data}

@api_router.delete("/users/{user_id}")
async def deactivate_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Deactivate a user - Super Admin only"""
    require_role(current_user, min_level=3)
    
    if user_id == current_user["_id"]:
        raise HTTPException(status_code=400, detail="Impossible de désactiver votre propre compte")
    
    existing = await db.users.find_one({"_id": ObjectId(user_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"is_active": False}})
    await audit_log(current_user, "DELETE", "user", user_id, existing.get("name", ""))
    return {"message": "Utilisateur désactivé"}

# ============ APPROVAL REQUESTS ============
@api_router.get("/approvals")
async def get_approvals(status: str = None, current_user: dict = Depends(get_current_user)):
    """List approval requests"""
    perms = get_permissions(current_user.get("role", "user"))
    query = {}
    if status:
        query["status"] = status
    # Non-super admins can only see their own requests
    if not perms.get("can_approve"):
        query["requester_id"] = current_user["_id"]
    
    result = []
    async for a in db.approval_requests.find(query).sort("created_at", -1):
        result.append({
            "id": str(a["_id"]),
            "requester_id": a.get("requester_id", ""),
            "requester_name": a.get("requester_name", ""),
            "requester_role": a.get("requester_role", ""),
            "action": a.get("action", ""),
            "entity_type": a.get("entity_type", ""),
            "entity_id": a.get("entity_id", ""),
            "entity_name": a.get("entity_name", ""),
            "details": a.get("details", {}),
            "status": a.get("status", "pending"),
            "reviewed_by": a.get("reviewed_by", ""),
            "reviewed_at": a.get("reviewed_at", ""),
            "created_at": a.get("created_at", "")
        })
    return result

@api_router.get("/approvals/count")
async def get_approval_count(current_user: dict = Depends(get_current_user)):
    """Get count of pending approvals"""
    perms = get_permissions(current_user.get("role", "user"))
    if not perms.get("can_approve"):
        return {"count": 0}
    count = await db.approval_requests.count_documents({"status": "pending"})
    return {"count": count}

@api_router.post("/approvals/{approval_id}/approve")
async def approve_request(approval_id: str, current_user: dict = Depends(get_current_user)):
    """Approve a pending request - Super Admin only"""
    require_role(current_user, min_level=3)
    
    approval = await db.approval_requests.find_one({"_id": ObjectId(approval_id), "status": "pending"})
    if not approval:
        raise HTTPException(status_code=404, detail="Demande non trouvée ou déjà traitée")
    
    # Execute the approved action
    action = approval.get("action", "")
    entity_type = approval.get("entity_type", "")
    entity_id = approval.get("entity_id", "")
    
    if action == "delete_client":
        existing = await db.clients.find_one({"_id": ObjectId(entity_id), **SOFT_DELETE_FILTER})
        if existing:
            appart_ids = existing.get("appartement_ids") or []
            if not appart_ids and existing.get("appartement_id"):
                appart_ids = [existing["appartement_id"]]
            for aid in appart_ids:
                try:
                    await db.appartements.update_one({"_id": ObjectId(aid)}, {"$set": {"statut": "disponible", "client_id": None}})
                except Exception:
                    pass
            await db.clients.update_one({"_id": ObjectId(entity_id)}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": current_user["_id"], "deleted_by_name": current_user.get("name", "")}})
            await audit_log(current_user, "DELETE", "client", entity_id, existing.get("nom", ""), new_values={"approved_from": approval_id})
    elif action == "delete_appartement":
        existing = await db.appartements.find_one({"_id": ObjectId(entity_id), **SOFT_DELETE_FILTER})
        if existing:
            await db.appartements.update_one({"_id": ObjectId(entity_id)}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": current_user["_id"]}})
            await audit_log(current_user, "DELETE", "appartement", entity_id, existing.get("numero_lot", ""), new_values={"approved_from": approval_id})
    elif action == "delete_prospect":
        existing = await db.prospects.find_one({"_id": ObjectId(entity_id), **SOFT_DELETE_FILTER})
        if existing:
            await db.prospects.update_one({"_id": ObjectId(entity_id)}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": current_user["_id"]}})
            await audit_log(current_user, "DELETE", "prospect", entity_id, existing.get("nom", ""), new_values={"approved_from": approval_id})
    
    # Mark as approved
    await db.approval_requests.update_one(
        {"_id": ObjectId(approval_id)},
        {"$set": {"status": "approved", "reviewed_by": current_user.get("name", current_user.get("email", "")), "reviewed_at": datetime.now(timezone.utc).isoformat()}}
    )
    await audit_log(current_user, "APPROVE", "approval", approval_id, approval.get("entity_name", ""))
    await manager.broadcast({"type": "approval_processed"})
    
    # Send email notification to requester
    try:
        requester = await db.users.find_one({"_id": ObjectId(approval.get("requester_id"))})
        if requester and requester.get("email"):
            await send_approval_email(requester["email"], approval, "approved", current_user.get("name", ""))
    except Exception:
        pass
    
    return {"message": "Demande approuvée et action exécutée"}

@api_router.post("/approvals/{approval_id}/reject")
async def reject_request(approval_id: str, current_user: dict = Depends(get_current_user)):
    """Reject a pending request - Super Admin only"""
    require_role(current_user, min_level=3)
    
    approval = await db.approval_requests.find_one({"_id": ObjectId(approval_id), "status": "pending"})
    if not approval:
        raise HTTPException(status_code=404, detail="Demande non trouvée ou déjà traitée")
    
    await db.approval_requests.update_one(
        {"_id": ObjectId(approval_id)},
        {"$set": {"status": "rejected", "reviewed_by": current_user.get("name", current_user.get("email", "")), "reviewed_at": datetime.now(timezone.utc).isoformat()}}
    )
    await audit_log(current_user, "REJECT", "approval", approval_id, approval.get("entity_name", ""))
    await manager.broadcast({"type": "approval_processed"})
    
    # Send email notification to requester
    try:
        requester = await db.users.find_one({"_id": ObjectId(approval.get("requester_id"))})
        if requester and requester.get("email"):
            await send_approval_email(requester["email"], approval, "rejected", current_user.get("name", ""))
    except Exception:
        pass
    
    return {"message": "Demande rejetée"}

async def send_approval_email(to_email: str, approval: dict, decision: str, reviewer_name: str):
    """Send approval decision email via Resend"""
    try:
        if not resend.api_key:
            return
        action_label = {"delete_client": "Suppression client", "delete_appartement": "Suppression appartement", "delete_prospect": "Suppression prospect", "modify_price": "Modification prix"}.get(approval.get("action", ""), approval.get("action", ""))
        status_label = "Approuvée" if decision == "approved" else "Rejetée"
        status_color = "#22c55e" if decision == "approved" else "#ef4444"
        
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #1E3A5F; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; font-size: 20px;">DJERBA CONSTRUCTION</h1>
            </div>
            <div style="padding: 20px; background: #f8f9fa;">
                <h2 style="color: {status_color};">Demande {status_label}</h2>
                <p><strong>Action:</strong> {action_label}</p>
                <p><strong>Élément:</strong> {approval.get('entity_name', '')}</p>
                <p><strong>Décision par:</strong> {reviewer_name}</p>
                <p style="color: #666; font-size: 12px; margin-top: 20px;">CRM DJERBA CONSTRUCTION - EDIMCO</p>
            </div>
        </div>
        """
        params = {"from": SENDER_EMAIL, "to": [to_email], "subject": f"Demande {status_label} - {action_label}", "html": html}
        await asyncio.to_thread(resend.Emails.send, params)
    except Exception as e:
        logger.error(f"Approval email error: {e}")

async def send_approval_notification_to_admins(approval_doc: dict):
    """Notify super admins about new approval request"""
    try:
        if not resend.api_key:
            return
        notification_email = os.environ.get("NOTIFICATION_EMAIL", "")
        if not notification_email:
            return
        
        action_label = {"delete_client": "Suppression client", "delete_appartement": "Suppression appartement", "delete_prospect": "Suppression prospect"}.get(approval_doc.get("action", ""), approval_doc.get("action", ""))
        
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #1E3A5F; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; font-size: 20px;">DJERBA CONSTRUCTION</h1>
            </div>
            <div style="padding: 20px; background: #f8f9fa;">
                <h2 style="color: #C41E3A;">Nouvelle demande d'approbation</h2>
                <p><strong>Demandeur:</strong> {approval_doc.get('requester_name', '')}</p>
                <p><strong>Action:</strong> {action_label}</p>
                <p><strong>Élément:</strong> {approval_doc.get('entity_name', '')}</p>
                <p><strong>Détails:</strong> {json.dumps(approval_doc.get('details', {}), ensure_ascii=False)}</p>
                <a href="{os.environ.get('FRONTEND_URL', '')}/admin" style="display: inline-block; background: #C41E3A; color: white; padding: 12px 24px; text-decoration: none; border-radius: 4px; margin-top: 16px;">Voir dans le CRM</a>
                <p style="color: #666; font-size: 12px; margin-top: 20px;">CRM DJERBA CONSTRUCTION - EDIMCO</p>
            </div>
        </div>
        """
        params = {"from": SENDER_EMAIL, "to": [notification_email], "subject": f"Approbation requise: {action_label} - {approval_doc.get('entity_name', '')}", "html": html}
        await asyncio.to_thread(resend.Emails.send, params)
    except Exception as e:
        logger.error(f"Admin notification email error: {e}")

# ============ DUPLICATE CLIENTS ============
@api_router.get("/clients/duplicates")
async def get_duplicate_clients(current_user: dict = Depends(get_current_user)):
    """Find all potential duplicate clients"""
    require_role(current_user, min_level=2)
    
    clients_list = []
    async for c in db.clients.find(SOFT_DELETE_FILTER):
        clients_list.append({
            "id": str(c["_id"]),
            "reference": c.get("reference", ""),
            "nom": c.get("nom", ""),
            "telephone": c.get("telephone", ""),
            "telephone2": c.get("telephone2", ""),
            "email": c.get("email", ""),
            "statut": c.get("statut", ""),
            "created_at": c.get("created_at", ""),
            "appartement_ids": c.get("appartement_ids", []),
        })
    
    # Find duplicates by phone or name
    groups = {}
    for c in clients_list:
        keys = []
        if c["telephone"]:
            keys.append(f"tel:{c['telephone']}")
        if c.get("telephone2"):
            keys.append(f"tel:{c['telephone2']}")
        if c["nom"]:
            keys.append(f"nom:{c['nom'].lower().strip()}")
        if c.get("email"):
            keys.append(f"email:{c['email'].lower().strip()}")
        
        for key in keys:
            if key not in groups:
                groups[key] = []
            groups[key].append(c)
    
    # Collect groups with >1 client
    seen_pairs = set()
    duplicate_groups = []
    for key, members in groups.items():
        if len(members) > 1:
            ids = tuple(sorted(m["id"] for m in members))
            if ids not in seen_pairs:
                seen_pairs.add(ids)
                reason = key.split(":")[0]
                duplicate_groups.append({
                    "reason": reason,
                    "match_value": key.split(":", 1)[1],
                    "clients": members
                })
    
    return {"groups": duplicate_groups, "total": len(duplicate_groups)}

@api_router.post("/clients/merge/{keep_id}/{merge_id}")
async def merge_clients_action(keep_id: str, merge_id: str, current_user: dict = Depends(get_current_user)):
    """Merge merge_id into keep_id"""
    require_role(current_user, min_level=3)
    
    keep = await db.clients.find_one({"_id": ObjectId(keep_id), **SOFT_DELETE_FILTER})
    merge = await db.clients.find_one({"_id": ObjectId(merge_id), **SOFT_DELETE_FILTER})
    
    if not keep or not merge:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # Merge data: keep existing values, fill blanks from merge
    update_fields = {}
    for field in ["telephone2", "email", "salaire", "budget_min", "budget_max", "objectif", "mode_paiement", "etage_souhaite", "situation_familiale"]:
        if not keep.get(field) and merge.get(field):
            update_fields[field] = merge[field]
    
    # Merge notes
    if merge.get("notes"):
        existing_notes = keep.get("notes", "") or ""
        update_fields["notes"] = f"{existing_notes}\n[Fusionné de {merge.get('reference', merge_id)}]: {merge['notes']}".strip()
    
    # Merge apartment IDs
    keep_apts = set(keep.get("appartement_ids") or [])
    merge_apts = set(merge.get("appartement_ids") or [])
    if merge.get("appartement_id"):
        merge_apts.add(merge["appartement_id"])
    combined_apts = list(keep_apts | merge_apts)
    update_fields["appartement_ids"] = combined_apts
    
    # Transfer apartments ownership
    for aid in merge_apts - keep_apts:
        try:
            await db.appartements.update_one({"_id": ObjectId(aid)}, {"$set": {"client_id": keep_id}})
        except Exception:
            pass
    
    # Update reservations
    await db.reservations.update_many({"client_id": merge_id}, {"$set": {"client_id": keep_id}})
    
    # Apply updates to kept client
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.clients.update_one({"_id": ObjectId(keep_id)}, {"$set": update_fields})
    
    # Soft delete merged client
    await db.clients.update_one({"_id": ObjectId(merge_id)}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": current_user["_id"], "deleted_by_name": f"Fusionné → {keep.get('reference', keep_id)}"}})
    
    await audit_log(current_user, "MERGE", "client", keep_id, keep.get("nom", ""), old_values={"merged_from": merge_id, "merged_reference": merge.get("reference", "")}, new_values=update_fields)
    await manager.broadcast({"type": "client_updated", "data": {"id": keep_id}})
    
    return {"message": f"Client fusionné avec succès", "kept_id": keep_id, "merged_id": merge_id}

# ============ BACKUP & RESTORE ============
@api_router.get("/backups")
async def get_backups(current_user: dict = Depends(get_current_user)):
    """List all backups - Super Admin only"""
    require_role(current_user, min_level=3)
    return backup_manager.list_backups()

@api_router.post("/backups")
async def create_backup_endpoint(current_user: dict = Depends(get_current_user)):
    """Create a manual backup - Super Admin only"""
    require_role(current_user, min_level=3)
    
    result = await backup_manager.create_backup(
        backup_type="manual",
        triggered_by=current_user.get("name", current_user.get("email", ""))
    )
    
    # Log the backup action
    await audit_log(current_user, "BACKUP", "system", result["backup_id"], f"Backup {result['type']}", new_values={"status": result["status"], "size_mb": result.get("size_mb", 0)})
    
    # Save backup metadata to DB
    result_db = {k: v for k, v in result.items() if k != "_id"}
    await db.backups.insert_one(result_db)
    
    if result["status"] == "failed":
        # Send failure email
        asyncio.create_task(send_backup_alert_email("failed", result))
        raise HTTPException(status_code=500, detail=f"Sauvegarde échouée: {result.get('error', 'Unknown')}")
    
    return result

@api_router.post("/backups/{backup_id}/restore")
async def restore_backup_endpoint(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Restore a backup - Super Admin only"""
    require_role(current_user, min_level=3)
    
    # Create a safety backup before restoring
    safety = await backup_manager.create_backup(
        backup_type="pre_restore",
        triggered_by=f"Auto-save before restore by {current_user.get('name', '')}"
    )
    safety_db = {k: v for k, v in safety.items() if k != "_id"}
    await db.backups.insert_one(safety_db)
    
    result = await backup_manager.restore_backup(backup_id)
    
    if result["status"] == "failed":
        asyncio.create_task(send_backup_alert_email("restore_failed", result))
        raise HTTPException(status_code=500, detail=f"Restauration échouée: {result.get('error', 'Unknown')}")
    
    # Log restore and send email
    await audit_log(current_user, "RESTORE", "system", backup_id, f"Restore from {backup_id}", new_values={"safety_backup": safety.get("backup_id", "")})
    asyncio.create_task(send_backup_alert_email("restored", {**result, "safety_backup": safety.get("backup_id", "")}))
    
    return {**result, "safety_backup_id": safety.get("backup_id", "")}

@api_router.delete("/backups/{backup_id}")
async def delete_backup_endpoint(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a backup - Super Admin only"""
    require_role(current_user, min_level=3)
    
    deleted = backup_manager.delete_backup(backup_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Sauvegarde non trouvée")
    
    await db.backups.delete_one({"backup_id": backup_id})
    await audit_log(current_user, "DELETE_BACKUP", "system", backup_id, f"Deleted backup {backup_id}")
    return {"message": "Sauvegarde supprimée"}

@api_router.get("/backups/{backup_id}/download")
async def download_backup(backup_id: str, current_user: dict = Depends(get_current_user)):
    """Download a backup as ZIP - Super Admin only"""
    require_role(current_user, min_level=3)
    
    backup_path = backup_manager.get_backup_path(backup_id)
    if not os.path.isdir(backup_path):
        raise HTTPException(status_code=404, detail="Sauvegarde non trouvée")
    
    import zipfile
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dirpath, dirnames, filenames in os.walk(backup_path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                arcname = os.path.relpath(filepath, backup_path)
                zf.write(filepath, arcname)
    
    zip_buffer.seek(0)
    await audit_log(current_user, "DOWNLOAD_BACKUP", "system", backup_id, f"Downloaded {backup_id}")
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={backup_id}.zip"}
    )

@api_router.get("/backups/email-settings")
async def get_backup_email_settings(current_user: dict = Depends(get_current_user)):
    """Get backup email export settings"""
    require_role(current_user, min_level=3)
    settings = await db.notification_settings.find_one({"type": "backup_email_export"})
    if not settings:
        return {"enabled": False, "email": os.environ.get("NOTIFICATION_EMAIL", ""), "schedule": "weekly"}
    return {"enabled": settings.get("enabled", False), "email": settings.get("email", ""), "schedule": settings.get("schedule", "weekly")}

class BackupEmailSettings(BaseModel):
    enabled: bool
    email: str

@api_router.post("/backups/email-settings")
async def save_backup_email_settings(settings: BackupEmailSettings, current_user: dict = Depends(get_current_user)):
    """Save backup email export settings"""
    require_role(current_user, min_level=3)
    await db.notification_settings.update_one(
        {"type": "backup_email_export"},
        {"$set": {"type": "backup_email_export", "enabled": settings.enabled, "email": settings.email, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.get("name", "")}},
        upsert=True
    )
    await audit_log(current_user, "UPDATE", "settings", "backup_email_export", "Email export settings", new_values={"enabled": settings.enabled, "email": settings.email})
    return {"enabled": settings.enabled, "email": settings.email}

@api_router.post("/backups/send-test-email")
async def send_test_backup_email(current_user: dict = Depends(get_current_user)):
    """Send a test backup email now"""
    require_role(current_user, min_level=3)
    settings = await db.notification_settings.find_one({"type": "backup_email_export"})
    if settings and settings.get("enabled"):
        await scheduled_weekly_email_export()
        return {"message": "Email de test envoyé"}
    raise HTTPException(status_code=400, detail="Activez d'abord l'export par email")

@api_router.get("/backups/stats")
async def get_backup_stats(current_user: dict = Depends(get_current_user)):
    """Get backup statistics"""
    require_role(current_user, min_level=3)
    backups = backup_manager.list_backups()
    
    total = len(backups)
    successful = len([b for b in backups if b.get("status") == "success"])
    failed = len([b for b in backups if b.get("status") == "failed"])
    total_size = sum(b.get("size_mb", 0) for b in backups if b.get("status") == "success")
    last_backup = backups[0] if backups else None
    
    auto_count = len([b for b in backups if b.get("type") in ("auto_6h", "auto_daily")])
    manual_count = len([b for b in backups if b.get("type") == "manual"])
    
    return {
        "total": total,
        "successful": successful,
        "failed": failed,
        "total_size_mb": round(total_size, 2),
        "last_backup": last_backup,
        "auto_count": auto_count,
        "manual_count": manual_count,
    }

async def send_backup_alert_email(event_type: str, details: dict):
    """Send backup alert email via Resend"""
    try:
        if not resend.api_key:
            return
        notification_email = os.environ.get("NOTIFICATION_EMAIL", "")
        if not notification_email:
            return
        
        subjects = {
            "failed": "ALERTE: Sauvegarde échouée",
            "restore_failed": "ALERTE: Restauration échouée",
            "restored": "INFO: Restauration effectuée",
        }
        colors_map = {
            "failed": "#ef4444",
            "restore_failed": "#ef4444",
            "restored": "#3b82f6",
        }
        
        subject = subjects.get(event_type, f"Backup: {event_type}")
        color = colors_map.get(event_type, "#6b7280")
        
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #1E3A5F; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; font-size: 20px;">DJERBA CONSTRUCTION - CRM</h1>
            </div>
            <div style="padding: 20px; background: #f8f9fa;">
                <h2 style="color: {color};">{subject}</h2>
                <p><strong>Type:</strong> {event_type}</p>
                <p><strong>Backup ID:</strong> {details.get('backup_id', '-')}</p>
                <p><strong>Date:</strong> {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}</p>
                {"<p><strong>Erreur:</strong> " + str(details.get('error', '')) + "</p>" if details.get('error') else ""}
                {"<p><strong>Sauvegarde de sécurité:</strong> " + str(details.get('safety_backup', '')) + "</p>" if details.get('safety_backup') else ""}
                <p style="color: #666; font-size: 12px; margin-top: 20px;">CRM DJERBA CONSTRUCTION - Système de sauvegarde automatique</p>
            </div>
        </div>
        """
        params = {"from": SENDER_EMAIL, "to": [notification_email], "subject": subject, "html": html}
        await asyncio.to_thread(resend.Emails.send, params)
    except Exception as e:
        logger.error(f"Backup alert email error: {e}")

# ============ BACKUP SCHEDULER ============
scheduler = AsyncIOScheduler()

async def scheduled_backup_6h():
    """Automatic backup every 6 hours"""
    try:
        result = await backup_manager.create_backup(backup_type="auto_6h", triggered_by="Scheduler")
        result_db = {k: v for k, v in result.items() if k != "_id"}
        await db.backups.insert_one(result_db)
        if result["status"] == "failed":
            await send_backup_alert_email("failed", result)
        # Apply retention policy
        await backup_manager.apply_retention_policy()
    except Exception as e:
        logger.error(f"Scheduled 6h backup error: {e}")

async def scheduled_backup_daily():
    """Daily backup at 2 AM"""
    try:
        result = await backup_manager.create_backup(backup_type="auto_daily", triggered_by="Scheduler")
        result_db = {k: v for k, v in result.items() if k != "_id"}
        await db.backups.insert_one(result_db)
        if result["status"] == "failed":
            await send_backup_alert_email("failed", result)
        await backup_manager.apply_retention_policy()
    except Exception as e:
        logger.error(f"Scheduled daily backup error: {e}")

async def scheduled_weekly_email_export():
    """Weekly email with ZIP backup attached - every Sunday at 3 AM"""
    try:
        # Check if feature is enabled
        settings = await db.notification_settings.find_one({"type": "backup_email_export"})
        if not settings or not settings.get("enabled", False):
            return
        
        recipient = settings.get("email") or os.environ.get("NOTIFICATION_EMAIL", "")
        if not recipient or not resend.api_key:
            return
        
        # Create a fresh backup
        result = await backup_manager.create_backup(backup_type="weekly_export", triggered_by="Email Export")
        result_db = {k: v for k, v in result.items() if k != "_id"}
        await db.backups.insert_one(result_db)
        
        if result["status"] != "success":
            await send_backup_alert_email("failed", result)
            return
        
        # Create ZIP
        import zipfile, base64
        backup_path = backup_manager.get_backup_path(result["backup_id"])
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for dirpath, dirnames, filenames in os.walk(backup_path):
                for filename in filenames:
                    filepath = os.path.join(dirpath, filename)
                    arcname = os.path.relpath(filepath, backup_path)
                    zf.write(filepath, arcname)
        zip_bytes = zip_buffer.getvalue()
        
        date_str = datetime.now(timezone.utc).strftime('%d/%m/%Y')
        size_kb = round(len(zip_bytes) / 1024, 1)
        
        # Get DB stats
        client_count = await db.clients.count_documents({"deleted_at": {"$eq": None}})
        appart_count = await db.appartements.count_documents({})
        prospect_count = await db.prospects.count_documents({"deleted_at": {"$eq": None}})
        
        html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #1E3A5F; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; font-size: 20px;">DJERBA CONSTRUCTION - CRM</h1>
            </div>
            <div style="padding: 20px; background: #f8f9fa;">
                <h2 style="color: #1E3A5F;">Sauvegarde hebdomadaire</h2>
                <p style="color: #333;">Voici votre copie de sécurité automatique du <strong>{date_str}</strong>.</p>
                <div style="background: white; border-radius: 8px; padding: 16px; margin: 16px 0; border: 1px solid #e2e8f0;">
                    <p style="margin: 4px 0;"><strong>Clients :</strong> {client_count}</p>
                    <p style="margin: 4px 0;"><strong>Appartements :</strong> {appart_count}</p>
                    <p style="margin: 4px 0;"><strong>Prospects :</strong> {prospect_count}</p>
                    <p style="margin: 4px 0;"><strong>Taille :</strong> {size_kb} KB</p>
                </div>
                <p style="color: #666; font-size: 13px;">Le fichier ZIP est joint à cet email. Conservez-le dans un endroit sûr.</p>
                <p style="color: #999; font-size: 11px; margin-top: 20px;">CRM DJERBA CONSTRUCTION - Export automatique hebdomadaire</p>
            </div>
        </div>
        """
        
        zip_b64 = base64.b64encode(zip_bytes).decode()
        params = {
            "from": SENDER_EMAIL,
            "to": [recipient],
            "subject": f"CRM EDIMCO - Sauvegarde hebdomadaire {date_str}",
            "html": html,
            "attachments": [{"filename": f"backup_crm_{datetime.now(timezone.utc).strftime('%Y%m%d')}.zip", "content": zip_b64}]
        }
        await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Weekly backup email sent to {recipient} ({size_kb} KB)")
        
    except Exception as e:
        logger.error(f"Weekly email export error: {e}")

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            await manager.broadcast({"type": "message", "data": data})
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# ============ CHANGE PASSWORD ============
class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

@api_router.put("/auth/change-password")
async def change_password(req: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"_id": ObjectId(current_user["_id"])})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    if not verify_password(req.current_password, user.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Mot de passe actuel incorrect")
    
    if len(req.new_password) < 6:
        raise HTTPException(status_code=400, detail="Le nouveau mot de passe doit contenir au moins 6 caractères")
    
    new_hash = hash_password(req.new_password)
    await db.users.update_one(
        {"_id": ObjectId(current_user["_id"])},
        {"$set": {"password_hash": new_hash}}
    )
    
    await audit_log(current_user, "UPDATE", "session", current_user["_id"], current_user.get("name", ""), new_values={"password": "***modifié***"})
    return {"message": "Mot de passe modifié avec succès"}

# ============ AUDIT LOG ROUTES ============
@api_router.get("/audit-logs")
async def get_audit_logs(
    action: str = None, entity_type: str = None, user_id: str = None,
    date_from: str = None, date_to: str = None, search: str = None,
    limit: int = 200,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if action:
        query["action"] = action
    if entity_type:
        query["entity_type"] = entity_type
    if user_id:
        query["user_id"] = user_id
    if date_from:
        query["timestamp"] = query.get("timestamp", {})
        query["timestamp"]["$gte"] = date_from
    if date_to:
        query["timestamp"] = query.get("timestamp", {})
        query["timestamp"]["$lte"] = date_to
    if search:
        query["$or"] = [
            {"entity_name": {"$regex": search, "$options": "i"}},
            {"user_name": {"$regex": search, "$options": "i"}},
        ]
    
    result = []
    async for log in db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit):
        result.append(log)
    return result

# ============ TRASH / CORBEILLE ROUTES ============
@api_router.get("/trash")
async def get_trash(current_user: dict = Depends(get_current_user)):
    """Get all soft-deleted items across collections."""
    result = []
    
    async for c in db.clients.find({"deleted_at": {"$ne": None}}).sort("deleted_at", -1):
        result.append({
            "id": str(c["_id"]),
            "entity_type": "client",
            "entity_name": c.get("nom", ""),
            "deleted_at": c.get("deleted_at", ""),
            "deleted_by": c.get("deleted_by", ""),
            "deleted_by_name": c.get("deleted_by_name", ""),
            "data": {k: v for k, v in c.items() if k not in ("_id", "deleted_at", "deleted_by", "deleted_by_name")}
        })
    
    async for p in db.prospects.find({"deleted_at": {"$ne": None}}).sort("deleted_at", -1):
        result.append({
            "id": str(p["_id"]),
            "entity_type": "prospect",
            "entity_name": p.get("nom", ""),
            "deleted_at": p.get("deleted_at", ""),
            "deleted_by": p.get("deleted_by", ""),
            "deleted_by_name": p.get("deleted_by_name", ""),
            "data": {k: v for k, v in p.items() if k not in ("_id", "deleted_at", "deleted_by", "deleted_by_name")}
        })
    
    async for a in db.appartements.find({"deleted_at": {"$ne": None}}).sort("deleted_at", -1):
        result.append({
            "id": str(a["_id"]),
            "entity_type": "appartement",
            "entity_name": f"Lot {a.get('numero_lot', '')} Bloc {a.get('bloc', '')}",
            "deleted_at": a.get("deleted_at", ""),
            "deleted_by": a.get("deleted_by", ""),
            "deleted_by_name": a.get("deleted_by_name", ""),
            "data": {k: v for k, v in a.items() if k not in ("_id", "deleted_at", "deleted_by", "deleted_by_name")}
        })
    
    result.sort(key=lambda x: x.get("deleted_at", ""), reverse=True)
    return result

@api_router.post("/trash/{entity_type}/{entity_id}/restore")
async def restore_from_trash(entity_type: str, entity_id: str, current_user: dict = Depends(get_current_user)):
    """Restore a soft-deleted item."""
    collection_map = {"client": "clients", "prospect": "prospects", "appartement": "appartements"}
    collection_name = collection_map.get(entity_type)
    if not collection_name:
        raise HTTPException(status_code=400, detail="Type d'entité invalide")
    
    collection = db[collection_name]
    existing = await collection.find_one({"_id": ObjectId(entity_id), "deleted_at": {"$ne": None}})
    if not existing:
        raise HTTPException(status_code=404, detail="Élément non trouvé dans la corbeille")
    
    await collection.update_one(
        {"_id": ObjectId(entity_id)},
        {"$set": {"deleted_at": None, "deleted_by": None, "deleted_by_name": None}}
    )
    
    await audit_log(current_user, "RESTORE", entity_type, entity_id, existing.get("nom", existing.get("numero_lot", "")))
    await manager.broadcast({"type": f"{entity_type}_restored", "data": {"id": entity_id}})
    return {"message": "Élément restauré avec succès"}

@api_router.delete("/trash/{entity_type}/{entity_id}/permanent")
async def permanent_delete(entity_type: str, entity_id: str, current_user: dict = Depends(get_current_user)):
    """Permanently delete an item - SUPER ADMIN ONLY."""
    require_role(current_user, min_level=3)
    
    collection_map = {"client": "clients", "prospect": "prospects", "appartement": "appartements"}
    collection_name = collection_map.get(entity_type)
    if not collection_name:
        raise HTTPException(status_code=400, detail="Type d'entité invalide")
    
    collection = db[collection_name]
    existing = await collection.find_one({"_id": ObjectId(entity_id), "deleted_at": {"$ne": None}})
    if not existing:
        raise HTTPException(status_code=404, detail="Élément non trouvé dans la corbeille")
    
    await collection.delete_one({"_id": ObjectId(entity_id)})
    
    await audit_log(current_user, "PERMANENT_DELETE", entity_type, entity_id, existing.get("nom", existing.get("numero_lot", "")))
    return {"message": "Élément supprimé définitivement"}

# ============ ADMIN SEED ENDPOINT ============
@api_router.post("/admin/seed")
async def admin_seed_edimco(current_user: dict = Depends(get_current_user)):
    """Force seed EDIMCO apartments - super admin only"""
    require_role(current_user, min_level=3)
    
    apparts_count = await db.appartements.count_documents({})
    if apparts_count > 0:
        return {"message": f"Base déjà peuplée: {apparts_count} appartements", "seeded": False}
    
    # Ensure EDIMCO residence exists
    residence = await db.residences.find_one({"nom": "EDIMCO"})
    if not residence:
        result = await db.residences.insert_one({
            "nom": "EDIMCO",
            "adresse": "EDIMCO, Commune de Bejaia, Wilaya de Bejaia",
            "description": "Résidence DJERBA - 264 logements promotionnels",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        residence = await db.residences.find_one({"_id": result.inserted_id})
    
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from seed_edimco import ALL_LOTS, PRIX_M2
    
    rid = str(residence["_id"])
    docs = []
    for lot in ALL_LOTS:
        docs.append({
            "residence_id": rid,
            "numero_lot": lot["lot"],
            "bloc": lot["bloc"],
            "etage": lot["etage"],
            "destination": lot["dest"],
            "type_appart": lot["type"],
            "surface_habitable": lot["sh"],
            "surface_utile": lot["su"],
            "prix": round(lot["sh"] * PRIX_M2),
            "statut": "disponible",
            "client_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    await db.appartements.insert_many(docs)
    await manager.broadcast({"type": "appartement_created"})
    return {"message": f"{len(docs)} lots EDIMCO créés", "seeded": True, "count": len(docs)}

# Include router
app.include_router(api_router)

# Health check endpoint (no auth required)
@app.get("/api/health")
async def health_check():
    try:
        # Test DB connection
        await db.command("ping")
        users_count = await db.users.count_documents({})
        clients_count = await db.clients.count_documents(SOFT_DELETE_FILTER)
        apparts_count = await db.appartements.count_documents(SOFT_DELETE_FILTER)
        return {
            "status": "ok",
            "database": "connected",
            "users": users_count,
            "clients": clients_count,
            "appartements": apparts_count
        }
    except Exception as e:
        return {"status": "error", "database": str(e)}

# Download build ZIP (no auth - temporary)
from fastapi.responses import FileResponse
@app.get("/api/download-build")
async def download_build():
    zip_path = "/app/crm-cpanel-build.zip"
    if os.path.exists(zip_path):
        return FileResponse(zip_path, filename="crm-cpanel-build.zip", media_type="application/zip")
    return {"error": "Build not found"}

# Root endpoint
@api_router.get("/")
async def root():
    return {"message": "DJERBA CONSTRUCTION CRM API", "version": "2.0"}

# Startup event
@app.on_event("startup")
async def startup_event():
    logger.info("=== STARTUP ===")
    logger.info(f"DB_NAME: {os.environ.get('DB_NAME')}")
    logger.info(f"MONGO_URL: {os.environ.get('MONGO_URL', '')[:40]}...")
    
    try:
        await db.command("ping")
        logger.info("MongoDB connection: OK")
    except Exception as e:
        logger.error(f"MongoDB connection FAILED: {e}")
        return
    
    await db.users.create_index("email", unique=True)
    
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@immo.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    
    # Also migrate existing "admin" role to "super_admin"
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "Administrateur",
            "role": "super_admin",
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin user created: {admin_email}")
    else:
        if existing.get("role") == "admin":
            await db.users.update_one({"email": admin_email}, {"$set": {"role": "super_admin"}})
            logger.info(f"Migrated admin role to super_admin: {admin_email}")
        if not verify_password(admin_password, existing["password_hash"]):
            await db.users.update_one(
                {"email": admin_email},
                {"$set": {"password_hash": hash_password(admin_password)}}
            )
            logger.info(f"Admin password updated: {admin_email}")
    
    # Seed EDIMCO residence
    try:
        residences_count = await db.residences.count_documents({})
        if residences_count == 0:
            await db.residences.insert_one(
                {"nom": "EDIMCO", "adresse": "EDIMCO, Commune de Bejaia, Wilaya de Bejaia", "description": "Résidence DJERBA - 264 logements promotionnels en R+11 avec sous-sol, duplex 10e-11e étage, RDC et 1er étage commercial/service, crèche et parking souterrain", "created_at": datetime.now(timezone.utc).isoformat()}
            )
            logger.info("EDIMCO residence created")
    except Exception as e:
        logger.error(f"Residence seed error: {e}")
    
    # Seed EDIMCO apartments if empty
    try:
        apparts_count = await db.appartements.count_documents({})
        if apparts_count == 0:
            logger.info("No apartments found — seeding EDIMCO lots...")
            residence = await db.residences.find_one({"nom": "EDIMCO"})
            if residence:
                import sys
                sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
                from seed_edimco import ALL_LOTS, PRIX_M2
                rid = str(residence["_id"])
                docs = []
                for lot in ALL_LOTS:
                    docs.append({
                        "residence_id": rid,
                        "numero_lot": lot["lot"],
                        "bloc": lot["bloc"],
                        "etage": lot["etage"],
                        "destination": lot["dest"],
                        "type_appart": lot["type"],
                        "surface_habitable": lot["sh"],
                        "surface_utile": lot["su"],
                        "prix": round(lot["sh"] * PRIX_M2),
                        "statut": "disponible",
                        "client_id": None,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    })
                await db.appartements.insert_many(docs)
                logger.info(f"Seeded {len(docs)} EDIMCO lots")
    except Exception as e:
        logger.error(f"Apartments seed error: {e}")
    
    # Write test credentials (non-critical, skip if filesystem is read-only)
    try:
        os.makedirs("/app/memory", exist_ok=True)
        with open("/app/memory/test_credentials.md", "w") as f:
            f.write(f"# Test Credentials\n\n## Admin Account\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: super_admin\n")
    except Exception:
        pass  # Not critical for production
    
    # ===== DATA MIGRATION =====
    # 1. Assign auto-references to clients that don't have one
    try:
        last_ref = await db.clients.find_one(
            {"reference": {"$exists": True, "$ne": "", "$ne": None}},
            sort=[("reference", -1)]
        )
        if last_ref and last_ref.get("reference"):
            try:
                counter = int(last_ref["reference"].replace("#", ""))
            except (ValueError, TypeError):
                counter = 0
        else:
            counter = 0
        
        clients_without_ref = db.clients.find({
            "$or": [
                {"reference": {"$exists": False}},
                {"reference": None},
                {"reference": ""}
            ]
        }).sort("created_at", 1)
        
        async for c in clients_without_ref:
            counter += 1
            ref = f"#{counter:03d}"
            await db.clients.update_one({"_id": c["_id"]}, {"$set": {"reference": ref}})
        
        if counter > 0:
            logger.info(f"Migration: assigned references up to #{counter:03d}")
    except Exception as e:
        logger.error(f"Reference migration error: {e}")
    
    # 2. Migrate old appartement_id (singular) to appartement_ids (array)
    try:
        migrated = 0
        async for c in db.clients.find({"appartement_id": {"$exists": True, "$ne": None}, "appartement_ids": {"$exists": False}}):
            old_id = c.get("appartement_id")
            if old_id:
                await db.clients.update_one({"_id": c["_id"]}, {"$set": {"appartement_ids": [old_id]}})
                migrated += 1
        # Also handle clients with old field but empty appartement_ids
        async for c in db.clients.find({"appartement_id": {"$exists": True, "$ne": None}, "appartement_ids": {"$size": 0}}):
            old_id = c.get("appartement_id")
            if old_id:
                await db.clients.update_one({"_id": c["_id"]}, {"$set": {"appartement_ids": [old_id]}})
                migrated += 1
        if migrated > 0:
            logger.info(f"Migration: migrated {migrated} clients from appartement_id to appartement_ids")
    except Exception as e:
        logger.error(f"Apartment migration error: {e}")
    
    # 3. Migrate old 'admin' role to 'super_admin' for all users
    try:
        result = await db.users.update_many({"role": "admin"}, {"$set": {"role": "super_admin"}})
        if result.modified_count > 0:
            logger.info(f"Migration: migrated {result.modified_count} admin users to super_admin")
    except Exception as e:
        logger.error(f"Role migration error: {e}")
    
    # ===== START BACKUP SCHEDULER =====
    try:
        scheduler.add_job(scheduled_backup_6h, 'interval', hours=6, id='backup_6h', replace_existing=True)
        scheduler.add_job(scheduled_backup_daily, 'cron', hour=2, minute=0, id='backup_daily', replace_existing=True)
        scheduler.add_job(scheduled_weekly_email_export, 'cron', day_of_week='sun', hour=3, minute=0, id='weekly_email_export', replace_existing=True)
        scheduler.start()
        logger.info("Backup scheduler started (every 6h + daily at 02:00)")
    except Exception as e:
        logger.error(f"Scheduler start error: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        scheduler.shutdown(wait=False)
    except Exception:
        pass
    client.close()
